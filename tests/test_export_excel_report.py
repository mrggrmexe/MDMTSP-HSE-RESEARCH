from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

import export_excel_report as report


class ExportExcelReportTests(unittest.TestCase):
    def make_run_json(
        self,
        path: Path,
        *,
        run_id: str,
        algorithm_id: str,
        instance_name: str,
        objective: float,
        wall_time_ms: float,
        feasible: bool = True,
        status: str = "ok",
        instance_type: str = "random",
    ) -> None:
        payload = {
            "schema_version": 2,
            "run_id": run_id,
            "timestamp_utc": "2026-05-07T00:00:00Z",
            "algorithm_id": algorithm_id,
            "suite_name": "research_suite",
            "instance_name": instance_name,
            "instance_type": instance_type,
            "instance_path": f"/repo/instances/{instance_name}.json",
            "seed": 42,
            "depot_count": 2,
            "customer_count": 10,
            "salesman_count": 3,
            "return_to_depot": True,
            "objective": objective,
            "feasible": feasible,
            "status": status,
            "route_count": 3,
            "wall_time_ms": wall_time_ms,
            "wall_time_us": int(round(wall_time_ms * 1000.0)),
            "wall_time_s": wall_time_ms / 1000.0,
        }
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload), encoding="utf-8")

    def test_build_export_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            runs_root = root / "results" / "runs"
            logs_root = root / "results" / "logs"

            self.make_run_json(
                runs_root / "suite" / "nearest_neighbour" / "instance_a" / "run_nn.json",
                run_id="run_nn",
                algorithm_id="nearest_neighbour",
                instance_name="instance_a_random",
                objective=110.0,
                wall_time_ms=5.0,
            )
            self.make_run_json(
                runs_root / "suite" / "random_insertion" / "instance_a" / "run_ri.json",
                run_id="run_ri",
                algorithm_id="random_insertion",
                instance_name="instance_a_random",
                objective=100.0,
                wall_time_ms=8.0,
            )
            self.make_run_json(
                runs_root / "suite" / "nearest_neighbour" / "instance_b" / "run_fail.json",
                run_id="run_fail",
                algorithm_id="nearest_neighbour",
                instance_name="instance_b_grid",
                objective=200.0,
                wall_time_ms=6.0,
                feasible=False,
            )

            (logs_root / "suite" / "nearest_neighbour" / "instance_a" / "run_nn.log").parent.mkdir(parents=True, exist_ok=True)
            (logs_root / "suite" / "nearest_neighbour" / "instance_a" / "run_nn.log").write_text("warning: slow move\n", encoding="utf-8")
            (logs_root / "suite" / "random_insertion" / "instance_a" / "run_ri.log").parent.mkdir(parents=True, exist_ok=True)
            (logs_root / "suite" / "random_insertion" / "instance_a" / "run_ri.log").write_text("ok\n", encoding="utf-8")
            (logs_root / "suite" / "nearest_neighbour" / "instance_b" / "run_fail.log").parent.mkdir(parents=True, exist_ok=True)
            (logs_root / "suite" / "nearest_neighbour" / "instance_b" / "run_fail.log").write_text("error: infeasible solution\nTraceback\n", encoding="utf-8")

            bundle = report.build_export_bundle(
                [runs_root],
                [logs_root],
                root,
                max_tail_lines=4,
                max_cell_text=1000,
            )

            self.assertEqual(len(bundle.runs), 3)
            self.assertEqual(len(bundle.algorithms), 2)
            self.assertEqual(len(bundle.instances), 2)
            self.assertEqual(len(bundle.failures), 1)

            by_algorithm = {row["algorithm_id"]: row for row in bundle.algorithms}
            self.assertAlmostEqual(by_algorithm["random_insertion"]["median_gap_to_best_observed"], 0.0)
            self.assertGreater(by_algorithm["nearest_neighbour"]["median_gap_to_best_observed"], 0.0)

            baseline = {row["instance_name"]: row for row in bundle.baseline_matrix}
            self.assertEqual(baseline["instance_a_random"]["random_insertion__best_objective"], 100.0)
            self.assertEqual(baseline["instance_a_random"]["nearest_neighbour__best_objective"], 110.0)

    def test_write_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            runs_root = root / "results" / "runs"
            logs_root = root / "results" / "logs"
            self.make_run_json(
                runs_root / "suite" / "nearest_neighbour" / "instance_a" / "run_nn.json",
                run_id="run_nn",
                algorithm_id="nearest_neighbour",
                instance_name="instance_a_random",
                objective=110.0,
                wall_time_ms=5.0,
            )
            (logs_root / "suite" / "nearest_neighbour" / "instance_a" / "run_nn.log").parent.mkdir(parents=True, exist_ok=True)
            (logs_root / "suite" / "nearest_neighbour" / "instance_a" / "run_nn.log").write_text("ok\n", encoding="utf-8")

            bundle = report.build_export_bundle(
                [runs_root],
                [logs_root],
                root,
                max_tail_lines=4,
                max_cell_text=1000,
            )
            output_path = root / "report.xlsx"
            report.write_workbook(bundle, output_path, root)

            self.assertTrue(output_path.exists())
            workbook = openpyxl.load_workbook(output_path)
            self.assertIn("Overview", workbook.sheetnames)
            self.assertIn("Runs", workbook.sheetnames)
            self.assertIn("Algorithms", workbook.sheetnames)

    def test_collect_runs_records_parse_issue(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            broken = root / "results" / "runs" / "broken.json"
            broken.parent.mkdir(parents=True, exist_ok=True)
            broken.write_text("{not valid json", encoding="utf-8")
            rows, issues = report.collect_runs([broken.parent], root)
            self.assertEqual(rows, [])
            self.assertEqual(len(issues), 1)
            self.assertEqual(issues[0].source_kind, "run_json")


if __name__ == "__main__":
    unittest.main()
