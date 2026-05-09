#!/usr/bin/env python3

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

GROUP_LONG_PATTERN = re.compile(r"group_(\d+)")
GROUP_SHORT_PATTERN = re.compile(r"^g(\d+)([a-zA-Z]*)_")


def pick_first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for name in candidates:
        if name in df.columns:
            return name
    return None


def as_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def remove_sheet_if_exists(wb, title: str) -> None:
    if title in wb.sheetnames:
        wb.remove(wb[title])


def natural_group_key(value: str) -> tuple[int, int, str, str]:
    value = value or ""

    short_match = GROUP_SHORT_PATTERN.search(value)
    if short_match:
        number = int(short_match.group(1))
        short_suffix = short_match.group(2)

        short_suffix_map = {
            "": "",
            "h": "hard",
            "t": "traps",
        }
        suffix_token = short_suffix_map.get(short_suffix, short_suffix)

        suffix_priority = {
            "": 0,
            "hard": 1,
            "traps": 2,
            "trap": 2,
        }

        return (
            number,
            suffix_priority.get(suffix_token, 999),
            suffix_token,
            value,
        )

    long_match = GROUP_LONG_PATTERN.search(value)
    if long_match:
        number = int(long_match.group(1))

        suffix_token = ""
        if "_hard" in value:
            suffix_token = "hard"
        elif "_traps" in value or "_trap" in value:
            suffix_token = "traps"

        suffix_priority = {
            "": 0,
            "hard": 1,
            "traps": 2,
        }

        return (
            number,
            suffix_priority.get(suffix_token, 999),
            suffix_token,
            value,
        )

    return (10**9, 999, "", value)


def ordered_unique_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def read_known_instances(instance_summary_csv: Path | None) -> list[str]:
    if instance_summary_csv is None or not instance_summary_csv.exists():
        return []

    df = pd.read_csv(instance_summary_csv)
    if df.empty:
        return []

    instance_col = pick_first_existing_column(df, ["instance_name", "instance"])
    if instance_col is None:
        return []

    values = [str(x) for x in df[instance_col].dropna().tolist()]
    return sorted(ordered_unique_strings(values), key=natural_group_key)


def expand_matrix_columns(matrix: pd.DataFrame, known_instances: list[str]) -> pd.DataFrame:
    matrix = matrix.copy()
    matrix.columns = [str(c) for c in matrix.columns]
    matrix.index = [str(i) for i in matrix.index]

    ordered_rows = sorted(matrix.index.tolist())

    if not known_instances:
        ordered_columns = sorted(matrix.columns.tolist(), key=natural_group_key)
        return matrix.reindex(index=ordered_rows, columns=ordered_columns)

    known_order = ordered_unique_strings(known_instances)
    known_set = set(known_order)
    extras = [col for col in matrix.columns if col not in known_set]
    extras = sorted(extras, key=natural_group_key)

    ordered_columns = known_order + extras
    return matrix.reindex(index=ordered_rows, columns=ordered_columns)


def write_matrix_sheet(
    wb,
    title: str,
    matrix: pd.DataFrame,
    legend_text: str,
    number_format: str,
) -> None:
    remove_sheet_if_exists(wb, title)
    ws = wb.create_sheet(title)

    ws["A1"] = legend_text
    ws["A1"].font = Font(bold=True)

    ws.cell(row=2, column=1, value="algorithm_id")
    for j, col_name in enumerate(matrix.columns, start=2):
        ws.cell(row=2, column=j, value=str(col_name))

    for i, (row_name, row_values) in enumerate(matrix.iterrows(), start=3):
        ws.cell(row=i, column=1, value=str(row_name))
        for j, value in enumerate(row_values.tolist(), start=2):
            cell = ws.cell(row=i, column=j)
            cell.value = None if pd.isna(value) else float(value)

    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        cell.border = THIN_BORDER

    for row_idx in range(3, max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=2, max_col=max_col):
        for cell in row:
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5.5
    ws.row_dimensions[2].height = 140

    if max_row >= 3 and max_col >= 2:
        data_range = f"B3:{get_column_letter(max_col)}{max_row}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="percentile",
                start_value=10,
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="percentile",
                end_value=90,
                end_color="F8696B",
            ),
        )


def build_gap_matrix(df: pd.DataFrame, known_instances: list[str]) -> pd.DataFrame:
    algo_col = pick_first_existing_column(df, ["algorithm_id", "algorithm"])
    instance_col = pick_first_existing_column(df, ["instance_name", "instance"])
    value_col = pick_first_existing_column(
        df,
        [
            "median_gap_to_best_observed",
            "mean_gap_to_best_observed",
            "min_gap_to_best_observed",
        ],
    )
    if algo_col is None or instance_col is None or value_col is None:
        raise ValueError("algorithm_instance_summary.csv does not contain required gap heatmap columns")

    work = df[[algo_col, instance_col, value_col]].copy()
    work[value_col] = as_numeric(work[value_col])
    work[instance_col] = work[instance_col].astype(str)
    work = work.dropna(subset=[value_col])

    matrix = work.pivot_table(
        index=algo_col,
        columns=instance_col,
        values=value_col,
        aggfunc="median",
    )
    return expand_matrix_columns(matrix, known_instances)


def build_time_matrix(df: pd.DataFrame, known_instances: list[str]) -> tuple[pd.DataFrame, str]:
    algo_col = pick_first_existing_column(df, ["algorithm_id", "algorithm"])
    instance_col = pick_first_existing_column(df, ["instance_name", "instance"])
    value_col = pick_first_existing_column(
        df,
        [
            "median_time_ratio_to_fastest",
            "mean_time_ratio_to_fastest",
            "median_wall_time_ms",
            "mean_wall_time_ms",
        ],
    )
    if algo_col is None or instance_col is None or value_col is None:
        raise ValueError("algorithm_instance_summary.csv does not contain required time heatmap columns")

    work = df[[algo_col, instance_col, value_col]].copy()
    work[value_col] = as_numeric(work[value_col])
    work[instance_col] = work[instance_col].astype(str)
    work = work.dropna(subset=[value_col])

    matrix = work.pivot_table(
        index=algo_col,
        columns=instance_col,
        values=value_col,
        aggfunc="median",
    )
    return expand_matrix_columns(matrix, known_instances), value_col


def add_heatmap_sheets_to_workbook(
    workbook_path: Path,
    algorithm_instance_summary_csv: Path,
    instance_summary_csv: Path | None = None,
) -> None:
    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")
    if not algorithm_instance_summary_csv.exists():
        raise FileNotFoundError(f"CSV not found: {algorithm_instance_summary_csv}")

    df = pd.read_csv(algorithm_instance_summary_csv)
    if df.empty:
        raise ValueError(f"CSV is empty: {algorithm_instance_summary_csv}")

    known_instances = read_known_instances(instance_summary_csv)

    gap_matrix = build_gap_matrix(df, known_instances)
    time_matrix, time_metric_name = build_time_matrix(df, known_instances)

    wb = load_workbook(workbook_path)

    write_matrix_sheet(
        wb=wb,
        title="Gap_Heatmap",
        matrix=gap_matrix,
        legend_text="Median gap to best observed, % (lower is better). Blank cells mean missing run/result.",
        number_format="0.000",
    )
    write_matrix_sheet(
        wb=wb,
        title="Time_Heatmap",
        matrix=time_matrix,
        legend_text=f"{time_metric_name} (lower is better). Blank cells mean missing run/result.",
        number_format="0.000",
    )

    wb.save(workbook_path)
