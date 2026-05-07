from __future__ import annotations

import argparse
import json
import math
import re
import statistics
import sys
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path, PurePosixPath
from tempfile import NamedTemporaryFile
from typing import Any, Iterable, Iterator, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - exercised only in misconfigured envs
    raise SystemExit(
        "openpyxl is required for Excel export. Install it with 'pip install openpyxl'."
    ) from exc


SUCCESS_STATUSES = {"ok", "success"}
KNOWN_INSTANCE_TYPES = {"random", "clustered", "grid", "adversarial", "line"}
EPS = 1e-9
MAX_SHEET_NAME_LEN = 31
DEFAULT_MAX_TAIL_LINES = 8
DEFAULT_MAX_CELL_TEXT = 32767
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
SUCCESS_FILL = PatternFill("solid", fgColor="E2F0D9")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ERROR_PATTERNS = (
    re.compile(r"\b(error|fatal|exception|traceback|failed|failure|cannot|can't|unable)\b", re.IGNORECASE),
    re.compile(r"\bsegmentation fault\b", re.IGNORECASE),
)
WARNING_PATTERNS = (
    re.compile(r"\bwarning\b", re.IGNORECASE),
    re.compile(r"\bdeprecated\b", re.IGNORECASE),
)


@dataclass(slots=True)
class ParseIssue:
    source_kind: str
    path: str
    reason: str


@dataclass(slots=True)
class LogRecord:
    run_id: str
    log_file: str
    relative_log_path: str | None
    log_exists: bool
    size_bytes: int
    line_count: int
    warning_count: int
    error_count: int
    traceback_count: int
    has_warning: bool
    has_error: bool
    first_warning_line: str | None
    first_error_line: str | None
    last_nonempty_line: str | None
    tail_excerpt: str | None

    def as_row(self) -> dict[str, Any]:
        return {
            "run_id": self.run_id,
            "log_file": self.log_file,
            "relative_log_path": self.relative_log_path,
            "log_exists": self.log_exists,
            "size_bytes": self.size_bytes,
            "line_count": self.line_count,
            "warning_count": self.warning_count,
            "error_count": self.error_count,
            "traceback_count": self.traceback_count,
            "has_warning": self.has_warning,
            "has_error": self.has_error,
            "first_warning_line": self.first_warning_line,
            "first_error_line": self.first_error_line,
            "last_nonempty_line": self.last_nonempty_line,
            "tail_excerpt": self.tail_excerpt,
        }


@dataclass(slots=True)
class ExportBundle:
    runs: list[dict[str, Any]]
    algorithms: list[dict[str, Any]]
    instances: list[dict[str, Any]]
    algorithm_instance: list[dict[str, Any]]
    algorithm_instance_type: list[dict[str, Any]]
    baseline_matrix: list[dict[str, Any]]
    failures: list[dict[str, Any]]
    logs: list[dict[str, Any]]
    parse_issues: list[dict[str, Any]]


@dataclass(frozen=True, slots=True)
class SummaryContext:
    best_objective_by_instance: dict[str, float]
    fastest_wall_time_ms_by_instance: dict[str, float]


class ExportError(RuntimeError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a research-grade Excel report from MDMTSP run JSON files and logs."
    )
    parser.add_argument(
        "--runs-root",
        dest="runs_roots",
        type=Path,
        action="append",
        help="Directory with per-run JSON files. Can be passed multiple times.",
    )
    parser.add_argument(
        "--logs-root",
        dest="logs_roots",
        type=Path,
        action="append",
        help="Directory with per-run .log files. Can be passed multiple times.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("results/reports/research_report.xlsx"),
        help="Target XLSX file.",
    )
    parser.add_argument(
        "--max-tail-lines",
        type=int,
        default=DEFAULT_MAX_TAIL_LINES,
        help="How many trailing log lines to store in workbook cells.",
    )
    parser.add_argument(
        "--max-cell-text",
        type=int,
        default=4000,
        help="Maximum number of characters kept in large text cells.",
    )
    parser.add_argument(
        "--repo-root",
        type=Path,
        default=None,
        help="Optional repository root for prettier relative paths.",
    )
    return parser.parse_args()


def repo_root_from_script() -> Path:
    return Path(__file__).resolve().parent.parent


def resolve_paths(raw_paths: Sequence[Path] | None, fallback: Path) -> list[Path]:
    if raw_paths:
        return [path.resolve() for path in raw_paths]
    return [fallback.resolve()]


def iter_files_with_suffix(roots: Sequence[Path], suffix: str) -> Iterator[Path]:
    seen: set[Path] = set()
    for root in roots:
        if not root.exists():
            continue
        if root.is_file():
            candidates = [root] if root.suffix.lower() == suffix else []
        else:
            candidates = root.rglob(f"*{suffix}")
        for path in candidates:
            if not path.is_file():
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            yield resolved


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def dig(data: dict[str, Any], *keys: str) -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict) or key not in current:
            return None
        current = current[key]
    return current


def first_non_none(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def as_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return str(value)


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return int(round(value))
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        return int(round(float(text)))
    return int(value)


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        value_float = float(value)
        if math.isnan(value_float) or math.isinf(value_float):
            return None
        return value_float
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        value_float = float(text)
        if math.isnan(value_float) or math.isinf(value_float):
            return None
        return value_float
    value_float = float(value)
    if math.isnan(value_float) or math.isinf(value_float):
        return None
    return value_float


def as_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"true", "1", "yes", "y"}:
            return True
        if text in {"false", "0", "no", "n"}:
            return False
    return None


def safe_mean(values: Sequence[float]) -> float | None:
    return statistics.fmean(values) if values else None


def safe_median(values: Sequence[float]) -> float | None:
    return statistics.median(values) if values else None


def percentile(values: Sequence[float], q: float) -> float | None:
    if not values:
        return None
    if q <= 0.0:
        return min(values)
    if q >= 1.0:
        return max(values)
    ordered = sorted(values)
    index = (len(ordered) - 1) * q
    low = int(index)
    high = min(low + 1, len(ordered) - 1)
    frac = index - low
    return ordered[low] * (1.0 - frac) + ordered[high] * frac


def success_from_status(status: str | None) -> bool:
    if status is None:
        return False
    return status.strip().lower() in SUCCESS_STATUSES


def approx_equal(a: float, b: float) -> bool:
    scale = max(1.0, abs(a), abs(b))
    return abs(a - b) <= EPS * scale


def infer_instance_type(instance_name: str | None, path: Path) -> str | None:
    candidates: list[str] = []
    if instance_name:
        candidates.extend(instance_name.split("_"))
    candidates.extend(path.stem.split("_"))
    for token in reversed(candidates):
        if token in KNOWN_INSTANCE_TYPES:
            return token
    return None


def normalize_runtime(data: dict[str, Any]) -> tuple[int | None, float | None, float | None]:
    wall_time_us = as_int(first_non_none(data.get("wall_time_us"), dig(data, "execution", "wall_time_us")))
    wall_time_ms = as_float(first_non_none(data.get("wall_time_ms"), dig(data, "execution", "wall_time_ms")))
    wall_time_s = as_float(first_non_none(data.get("wall_time_s"), dig(data, "execution", "wall_time_s")))

    if wall_time_us is None:
        if wall_time_ms is not None:
            wall_time_us = int(round(wall_time_ms * 1000.0))
        elif wall_time_s is not None:
            wall_time_us = int(round(wall_time_s * 1_000_000.0))
    if wall_time_ms is None and wall_time_us is not None:
        wall_time_ms = wall_time_us / 1000.0
    if wall_time_s is None and wall_time_us is not None:
        wall_time_s = wall_time_us / 1_000_000.0
    return wall_time_us, wall_time_ms, wall_time_s


def to_repo_relative(path_str: str | None, repo_root: Path | None) -> str | None:
    if not path_str:
        return None
    path = Path(path_str)
    if repo_root is None:
        return path_str
    try:
        return path.resolve().relative_to(repo_root.resolve()).as_posix()
    except Exception:
        return path_str


def normalize_run_json(path: Path, repo_root: Path | None) -> dict[str, Any]:
    data = load_json(path)
    if not isinstance(data, dict):
        raise ExportError("result JSON must be an object")

    routes = data.get("routes")
    route_count = len(routes) if isinstance(routes, list) else None
    wall_time_us, wall_time_ms, wall_time_s = normalize_runtime(data)

    instance_name = as_str(first_non_none(data.get("instance_name"), dig(data, "instance", "name")))
    if not instance_name:
        raise ExportError("missing instance_name")

    instance_path = as_str(first_non_none(data.get("instance_path"), dig(data, "instance", "path"), dig(data, "source", "instance_path")))
    customer_count = as_int(first_non_none(data.get("customer_count"), dig(data, "instance", "customer_count")))

    row: dict[str, Any] = {
        "run_id": as_str(first_non_none(data.get("run_id"), path.stem)) or path.stem,
        "timestamp_utc": as_str(first_non_none(data.get("timestamp_utc"), dig(data, "execution", "timestamp_utc"))),
        "suite_name": as_str(data.get("suite_name")),
        "algorithm_id": as_str(first_non_none(data.get("algorithm_id"), dig(data, "algorithm", "id"))) or "unknown",
        "instance_name": instance_name,
        "instance_type": as_str(first_non_none(data.get("instance_type"), dig(data, "instance", "instance_type"), infer_instance_type(instance_name, path))),
        "instance_path": instance_path,
        "relative_instance_path": to_repo_relative(instance_path, repo_root),
        "seed": as_int(first_non_none(data.get("seed"), dig(data, "execution", "seed"))),
        "improve_iterations": as_int(first_non_none(data.get("improve_iterations"), dig(data, "algorithm", "parameters", "improve_iterations"))),
        "depot_count": as_int(first_non_none(data.get("depot_count"), dig(data, "instance", "depot_count"))),
        "customer_count": customer_count,
        "salesman_count": as_int(first_non_none(data.get("salesman_count"), dig(data, "instance", "salesman_count"))),
        "return_to_depot": as_bool(first_non_none(data.get("return_to_depot"), dig(data, "instance", "return_to_depot"))),
        "objective": as_float(first_non_none(data.get("objective"), dig(data, "result", "objective"))),
        "feasible": as_bool(first_non_none(data.get("feasible"), dig(data, "result", "feasible"))),
        "status": as_str(first_non_none(data.get("status"), dig(data, "result", "status"))) or "unknown",
        "route_count": as_int(first_non_none(data.get("route_count"), dig(data, "result", "route_count"), route_count)),
        "wall_time_us": wall_time_us,
        "wall_time_ms": wall_time_ms,
        "wall_time_s": wall_time_s,
        "wall_time_per_customer_us": (
            wall_time_us / customer_count
            if wall_time_us is not None and customer_count is not None and customer_count > 0
            else None
        ),
        "result_file": str(path),
        "relative_result_file": to_repo_relative(str(path), repo_root),
    }
    return row


def summarize_log_text(text: str, *, max_tail_lines: int, max_cell_text: int) -> tuple[int, int, int, str | None, str | None, str | None, str | None, int]:
    lines = text.splitlines()
    nonempty_lines = [line.strip() for line in lines if line.strip()]
    warning_lines: list[str] = []
    error_lines: list[str] = []
    traceback_count = 0

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if any(pattern.search(line) for pattern in WARNING_PATTERNS):
            warning_lines.append(line)
        if any(pattern.search(line) for pattern in ERROR_PATTERNS):
            error_lines.append(line)
        if "traceback" in line.lower():
            traceback_count += 1

    tail_lines = nonempty_lines[-max(1, max_tail_lines):]
    tail_excerpt = "\n".join(tail_lines) if tail_lines else None
    tail_excerpt = trim_text(tail_excerpt, max_cell_text)

    return (
        len(lines),
        len(warning_lines),
        len(error_lines),
        trim_text(warning_lines[0], max_cell_text) if warning_lines else None,
        trim_text(error_lines[0], max_cell_text) if error_lines else None,
        trim_text(nonempty_lines[-1], max_cell_text) if nonempty_lines else None,
        tail_excerpt,
        traceback_count,
    )


def parse_log_file(path: Path, repo_root: Path | None, *, max_tail_lines: int, max_cell_text: int) -> LogRecord:
    try:
        text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = path.read_text(encoding="utf-8", errors="replace")

    line_count, warning_count, error_count, first_warning_line, first_error_line, last_nonempty_line, tail_excerpt, traceback_count = summarize_log_text(
        text,
        max_tail_lines=max_tail_lines,
        max_cell_text=max_cell_text,
    )

    return LogRecord(
        run_id=path.stem,
        log_file=str(path),
        relative_log_path=to_repo_relative(str(path), repo_root),
        log_exists=True,
        size_bytes=path.stat().st_size,
        line_count=line_count,
        warning_count=warning_count,
        error_count=error_count,
        traceback_count=traceback_count,
        has_warning=warning_count > 0,
        has_error=error_count > 0 or traceback_count > 0,
        first_warning_line=first_warning_line,
        first_error_line=first_error_line,
        last_nonempty_line=last_nonempty_line,
        tail_excerpt=tail_excerpt,
    )


def trim_text(value: str | None, max_len: int) -> str | None:
    if value is None:
        return None
    if max_len <= 0:
        return None
    if len(value) <= max_len:
        return value
    if max_len <= 1:
        return value[:max_len]
    return value[: max_len - 1] + "…"


def collect_runs(
    runs_roots: Sequence[Path],
    repo_root: Path | None,
) -> tuple[list[dict[str, Any]], list[ParseIssue]]:
    rows: list[dict[str, Any]] = []
    issues: list[ParseIssue] = []

    for path in iter_files_with_suffix(runs_roots, ".json"):
        try:
            row = normalize_run_json(path, repo_root)
        except Exception as exc:
            issues.append(ParseIssue("run_json", str(path), str(exc)))
            continue
        rows.append(row)

    rows.sort(key=lambda row: (
        row.get("suite_name") or "",
        row.get("algorithm_id") or "",
        row.get("instance_name") or "",
        row.get("seed") if row.get("seed") is not None else -1,
        row.get("run_id") or "",
    ))
    return rows, issues


def collect_logs(
    logs_roots: Sequence[Path],
    repo_root: Path | None,
    *,
    max_tail_lines: int,
    max_cell_text: int,
) -> tuple[dict[str, LogRecord], list[ParseIssue]]:
    records: dict[str, LogRecord] = {}
    issues: list[ParseIssue] = []

    for path in iter_files_with_suffix(logs_roots, ".log"):
        try:
            record = parse_log_file(
                path,
                repo_root,
                max_tail_lines=max_tail_lines,
                max_cell_text=max_cell_text,
            )
        except Exception as exc:
            issues.append(ParseIssue("log", str(path), str(exc)))
            continue
        records[record.run_id] = record

    return records, issues


def attach_logs(run_rows: list[dict[str, Any]], log_records: dict[str, LogRecord], repo_root: Path | None) -> None:
    for row in run_rows:
        log_record = log_records.get(str(row.get("run_id") or ""))
        if log_record is None:
            row.update(
                {
                    "log_file": None,
                    "relative_log_file": None,
                    "log_exists": False,
                    "log_size_bytes": None,
                    "log_line_count": None,
                    "log_warning_count": 0,
                    "log_error_count": 0,
                    "log_traceback_count": 0,
                    "log_has_warning": False,
                    "log_has_error": False,
                    "log_first_warning_line": None,
                    "log_first_error_line": None,
                    "log_last_nonempty_line": None,
                    "log_tail_excerpt": None,
                }
            )
            continue

        row.update(
            {
                "log_file": log_record.log_file,
                "relative_log_file": log_record.relative_log_path,
                "log_exists": log_record.log_exists,
                "log_size_bytes": log_record.size_bytes,
                "log_line_count": log_record.line_count,
                "log_warning_count": log_record.warning_count,
                "log_error_count": log_record.error_count,
                "log_traceback_count": log_record.traceback_count,
                "log_has_warning": log_record.has_warning,
                "log_has_error": log_record.has_error,
                "log_first_warning_line": log_record.first_warning_line,
                "log_first_error_line": log_record.first_error_line,
                "log_last_nonempty_line": log_record.last_nonempty_line,
                "log_tail_excerpt": log_record.tail_excerpt,
            }
        )

        row["result_file"] = str(Path(str(row["result_file"])).resolve())
        row["relative_result_file"] = to_repo_relative(row["result_file"], repo_root)


def build_summary_context(run_rows: Sequence[dict[str, Any]]) -> SummaryContext:
    feasible_rows = [row for row in run_rows if row.get("objective") is not None and success_from_status(as_str(row.get("status"))) and row.get("feasible") is not False]

    best_objective_by_instance: dict[str, float] = {}
    fastest_wall_time_ms_by_instance: dict[str, float] = {}

    for row in feasible_rows:
        instance_name = str(row["instance_name"])
        objective = as_float(row.get("objective"))
        wall_time_ms = as_float(row.get("wall_time_ms"))
        if objective is not None:
            best = best_objective_by_instance.get(instance_name)
            if best is None or objective < best:
                best_objective_by_instance[instance_name] = objective
        if wall_time_ms is not None:
            fastest = fastest_wall_time_ms_by_instance.get(instance_name)
            if fastest is None or wall_time_ms < fastest:
                fastest_wall_time_ms_by_instance[instance_name] = wall_time_ms

    return SummaryContext(
        best_objective_by_instance=best_objective_by_instance,
        fastest_wall_time_ms_by_instance=fastest_wall_time_ms_by_instance,
    )


def enrich_runs_with_relative_metrics(run_rows: list[dict[str, Any]], context: SummaryContext) -> None:
    for row in run_rows:
        instance_name = str(row["instance_name"])
        objective = as_float(row.get("objective"))
        wall_time_ms = as_float(row.get("wall_time_ms"))
        best_objective = context.best_objective_by_instance.get(instance_name)
        fastest_wall_time_ms = context.fastest_wall_time_ms_by_instance.get(instance_name)

        row["best_observed_objective"] = best_objective
        row["fastest_observed_wall_time_ms"] = fastest_wall_time_ms

        if objective is not None and best_objective is not None and abs(best_objective) > EPS:
            row["gap_to_best_observed"] = (objective - best_objective) / best_objective
        else:
            row["gap_to_best_observed"] = None

        if wall_time_ms is not None and fastest_wall_time_ms is not None and fastest_wall_time_ms > EPS:
            row["time_ratio_to_fastest"] = wall_time_ms / fastest_wall_time_ms
            row["time_gap_to_fastest"] = (wall_time_ms - fastest_wall_time_ms) / fastest_wall_time_ms
        else:
            row["time_ratio_to_fastest"] = None
            row["time_gap_to_fastest"] = None

        row["is_success"] = success_from_status(as_str(row.get("status")))
        row["is_feasible_success"] = bool(row["is_success"] and row.get("feasible") is not False)
        row["has_runtime_issue"] = bool(row.get("log_has_error"))
        row["is_flagged_failure"] = bool(
            not row["is_success"]
            or row.get("feasible") is False
            or row.get("log_has_error")
        )


def summarize_algorithms(run_rows: Sequence[dict[str, Any]], context: SummaryContext) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in run_rows:
        grouped[str(row.get("algorithm_id") or "unknown")].append(row)

    output: list[dict[str, Any]] = []
    for algorithm_id, rows in sorted(grouped.items()):
        objective_rows = [row for row in rows if row.get("objective") is not None and row.get("is_feasible_success")]
        wall_times_ms = [as_float(row.get("wall_time_ms")) for row in rows if as_float(row.get("wall_time_ms")) is not None and row.get("is_success")]
        gaps = [as_float(row.get("gap_to_best_observed")) for row in objective_rows if as_float(row.get("gap_to_best_observed")) is not None]
        time_ratios = [as_float(row.get("time_ratio_to_fastest")) for row in rows if as_float(row.get("time_ratio_to_fastest")) is not None and row.get("is_success")]
        time_gaps = [as_float(row.get("time_gap_to_fastest")) for row in rows if as_float(row.get("time_gap_to_fastest")) is not None and row.get("is_success")]
        wall_per_customer_us = [as_float(row.get("wall_time_per_customer_us")) for row in rows if as_float(row.get("wall_time_per_customer_us")) is not None and row.get("is_success")]

        best_run_count = 0
        fastest_run_count = 0
        best_instances: set[str] = set()
        fastest_instances: set[str] = set()
        for row in objective_rows:
            instance_name = str(row["instance_name"])
            objective = as_float(row.get("objective"))
            best_objective = context.best_objective_by_instance.get(instance_name)
            if objective is not None and best_objective is not None and approx_equal(objective, best_objective):
                best_run_count += 1
                best_instances.add(instance_name)

        for row in rows:
            if not row.get("is_success"):
                continue
            instance_name = str(row["instance_name"])
            wall_time_ms = as_float(row.get("wall_time_ms"))
            fastest_wall_time_ms = context.fastest_wall_time_ms_by_instance.get(instance_name)
            if wall_time_ms is not None and fastest_wall_time_ms is not None and approx_equal(wall_time_ms, fastest_wall_time_ms):
                fastest_run_count += 1
                fastest_instances.add(instance_name)

        output.append(
            {
                "algorithm_id": algorithm_id,
                "runs": len(rows),
                "successful_runs": sum(1 for row in rows if row.get("is_success")),
                "feasible_runs": sum(1 for row in rows if row.get("is_feasible_success")),
                "feasible_rate": (
                    sum(1 for row in rows if row.get("is_feasible_success")) / len(rows)
                    if rows else None
                ),
                "objective_comparable_runs": len(objective_rows),
                "time_comparable_runs": sum(1 for row in rows if row.get("time_ratio_to_fastest") is not None),
                "unique_instances": len({str(row["instance_name"]) for row in rows}),
                "unique_instance_types": len({str(row["instance_type"]) for row in rows if row.get("instance_type")}),
                "instance_types": ",".join(sorted({str(row["instance_type"]) for row in rows if row.get("instance_type")})),
                "unique_suites": len({str(row["suite_name"]) for row in rows if row.get("suite_name")}),
                "suite_names": ",".join(sorted({str(row["suite_name"]) for row in rows if row.get("suite_name")})),
                "best_run_count": best_run_count,
                "best_instance_coverage": len(best_instances),
                "fastest_run_count": fastest_run_count,
                "fastest_instance_coverage": len(fastest_instances),
                "min_gap_to_best_observed": min(gaps) if gaps else None,
                "mean_gap_to_best_observed": safe_mean(gaps),
                "median_gap_to_best_observed": safe_median(gaps),
                "min_time_ratio_to_fastest": min(time_ratios) if time_ratios else None,
                "mean_time_ratio_to_fastest": safe_mean(time_ratios),
                "median_time_ratio_to_fastest": safe_median(time_ratios),
                "min_time_gap_to_fastest": min(time_gaps) if time_gaps else None,
                "mean_time_gap_to_fastest": safe_mean(time_gaps),
                "median_time_gap_to_fastest": safe_median(time_gaps),
                "total_wall_time_s": (sum(value for value in wall_times_ms if value is not None) / 1000.0) if wall_times_ms else None,
                "mean_wall_time_ms": safe_mean([value for value in wall_times_ms if value is not None]),
                "median_wall_time_ms": safe_median([value for value in wall_times_ms if value is not None]),
                "p90_wall_time_ms": percentile([value for value in wall_times_ms if value is not None], 0.90),
                "median_wall_time_per_customer_us": safe_median([value for value in wall_per_customer_us if value is not None]),
                "log_error_runs": sum(1 for row in rows if row.get("log_has_error")),
                "log_warning_runs": sum(1 for row in rows if row.get("log_has_warning")),
            }
        )

    return output


def summarize_instances(run_rows: Sequence[dict[str, Any]], context: SummaryContext) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in run_rows:
        grouped[str(row["instance_name"])].append(row)

    output: list[dict[str, Any]] = []
    for instance_name, rows in sorted(grouped.items(), key=lambda item: (as_int(item[1][0].get("customer_count")) or -1, item[0])):
        objective_rows = [row for row in rows if row.get("objective") is not None and row.get("is_feasible_success")]
        runtime_rows = [row for row in rows if row.get("wall_time_ms") is not None and row.get("is_success")]
        best_objective = context.best_objective_by_instance.get(instance_name)
        fastest_wall_time_ms = context.fastest_wall_time_ms_by_instance.get(instance_name)

        best_algorithms: list[str] = []
        if best_objective is not None:
            best_algorithms = sorted(
                {
                    str(row["algorithm_id"])
                    for row in objective_rows
                    if as_float(row.get("objective")) is not None and approx_equal(as_float(row["objective"]), best_objective)
                }
            )

        fastest_algorithms: list[str] = []
        if fastest_wall_time_ms is not None:
            fastest_algorithms = sorted(
                {
                    str(row["algorithm_id"])
                    for row in runtime_rows
                    if as_float(row.get("wall_time_ms")) is not None and approx_equal(as_float(row["wall_time_ms"]), fastest_wall_time_ms)
                }
            )

        wall_times = [as_float(row.get("wall_time_ms")) for row in runtime_rows if as_float(row.get("wall_time_ms")) is not None]

        output.append(
            {
                "instance_name": instance_name,
                "instance_type": as_str(rows[0].get("instance_type")),
                "runs": len(rows),
                "successful_runs": sum(1 for row in rows if row.get("is_success")),
                "feasible_runs": sum(1 for row in rows if row.get("is_feasible_success")),
                "feasible_rate": (
                    sum(1 for row in rows if row.get("is_feasible_success")) / len(rows)
                    if rows else None
                ),
                "algorithms_tested": len({str(row["algorithm_id"]) for row in rows}),
                "seeds_tested": len({as_int(row.get("seed")) for row in rows if as_int(row.get("seed")) is not None}),
                "best_observed_objective": best_objective,
                "best_algorithm_ids": ",".join(best_algorithms),
                "fastest_observed_wall_time_ms": fastest_wall_time_ms,
                "fastest_algorithm_ids": ",".join(fastest_algorithms),
                "depot_count": as_int(rows[0].get("depot_count")),
                "customer_count": as_int(rows[0].get("customer_count")),
                "salesman_count": as_int(rows[0].get("salesman_count")),
                "return_to_depot": as_bool(rows[0].get("return_to_depot")),
                "median_wall_time_ms": safe_median(wall_times),
                "p90_wall_time_ms": percentile(wall_times, 0.90),
                "suite_names": ",".join(sorted({str(row["suite_name"]) for row in rows if row.get("suite_name")})),
            }
        )

    return output


def summarize_algorithm_instance(run_rows: Sequence[dict[str, Any]], context: SummaryContext) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in run_rows:
        grouped[(str(row["algorithm_id"]), str(row["instance_name"]))].append(row)

    output: list[dict[str, Any]] = []
    for (algorithm_id, instance_name), rows in sorted(grouped.items(), key=lambda item: ((as_int(item[1][0].get("customer_count")) or -1), item[0][1], item[0][0])):
        objective_rows = [row for row in rows if row.get("objective") is not None and row.get("is_feasible_success")]
        gaps = [as_float(row.get("gap_to_best_observed")) for row in objective_rows if as_float(row.get("gap_to_best_observed")) is not None]
        time_ratios = [as_float(row.get("time_ratio_to_fastest")) for row in rows if as_float(row.get("time_ratio_to_fastest")) is not None and row.get("is_success")]
        time_gaps = [as_float(row.get("time_gap_to_fastest")) for row in rows if as_float(row.get("time_gap_to_fastest")) is not None and row.get("is_success")]
        wall_times_ms = [as_float(row.get("wall_time_ms")) for row in rows if as_float(row.get("wall_time_ms")) is not None and row.get("is_success")]
        wall_per_customer_us = [as_float(row.get("wall_time_per_customer_us")) for row in rows if as_float(row.get("wall_time_per_customer_us")) is not None and row.get("is_success")]
        objectives = [as_float(row.get("objective")) for row in objective_rows if as_float(row.get("objective")) is not None]

        output.append(
            {
                "algorithm_id": algorithm_id,
                "instance_name": instance_name,
                "instance_type": as_str(rows[0].get("instance_type")),
                "runs": len(rows),
                "successful_runs": sum(1 for row in rows if row.get("is_success")),
                "feasible_runs": sum(1 for row in rows if row.get("is_feasible_success")),
                "feasible_rate": (
                    sum(1 for row in rows if row.get("is_feasible_success")) / len(rows)
                    if rows else None
                ),
                "suite_names": ",".join(sorted({str(row["suite_name"]) for row in rows if row.get("suite_name")})),
                "depot_count": as_int(rows[0].get("depot_count")),
                "customer_count": as_int(rows[0].get("customer_count")),
                "salesman_count": as_int(rows[0].get("salesman_count")),
                "return_to_depot": as_bool(rows[0].get("return_to_depot")),
                "best_objective": min(objectives) if objectives else None,
                "mean_objective": safe_mean(objectives),
                "median_objective": safe_median(objectives),
                "min_gap_to_best_observed": min(gaps) if gaps else None,
                "mean_gap_to_best_observed": safe_mean(gaps),
                "median_gap_to_best_observed": safe_median(gaps),
                "min_time_ratio_to_fastest": min(time_ratios) if time_ratios else None,
                "mean_time_ratio_to_fastest": safe_mean(time_ratios),
                "median_time_ratio_to_fastest": safe_median(time_ratios),
                "min_time_gap_to_fastest": min(time_gaps) if time_gaps else None,
                "mean_time_gap_to_fastest": safe_mean(time_gaps),
                "median_time_gap_to_fastest": safe_median(time_gaps),
                "total_wall_time_s": (sum(value for value in wall_times_ms if value is not None) / 1000.0) if wall_times_ms else None,
                "mean_wall_time_ms": safe_mean([value for value in wall_times_ms if value is not None]),
                "median_wall_time_ms": safe_median([value for value in wall_times_ms if value is not None]),
                "p90_wall_time_ms": percentile([value for value in wall_times_ms if value is not None], 0.90),
                "median_wall_time_per_customer_us": safe_median([value for value in wall_per_customer_us if value is not None]),
                "log_error_runs": sum(1 for row in rows if row.get("log_has_error")),
                "log_warning_runs": sum(1 for row in rows if row.get("log_has_warning")),
            }
        )

    return output


def summarize_algorithm_instance_type(run_rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in run_rows:
        instance_type = as_str(row.get("instance_type"))
        if not instance_type:
            continue
        grouped[(str(row["algorithm_id"]), instance_type)].append(row)

    output: list[dict[str, Any]] = []
    for (algorithm_id, instance_type), rows in sorted(grouped.items()):
        objective_rows = [row for row in rows if row.get("objective") is not None and row.get("is_feasible_success")]
        gaps = [as_float(row.get("gap_to_best_observed")) for row in objective_rows if as_float(row.get("gap_to_best_observed")) is not None]
        time_gaps = [as_float(row.get("time_gap_to_fastest")) for row in rows if as_float(row.get("time_gap_to_fastest")) is not None and row.get("is_success")]
        wall_times_ms = [as_float(row.get("wall_time_ms")) for row in rows if as_float(row.get("wall_time_ms")) is not None and row.get("is_success")]
        wall_per_customer_us = [as_float(row.get("wall_time_per_customer_us")) for row in rows if as_float(row.get("wall_time_per_customer_us")) is not None and row.get("is_success")]

        output.append(
            {
                "algorithm_id": algorithm_id,
                "instance_type": instance_type,
                "runs": len(rows),
                "successful_runs": sum(1 for row in rows if row.get("is_success")),
                "feasible_runs": sum(1 for row in rows if row.get("is_feasible_success")),
                "unique_instances": len({str(row["instance_name"]) for row in rows}),
                "median_gap_to_best_observed": safe_median(gaps),
                "mean_gap_to_best_observed": safe_mean(gaps),
                "median_time_gap_to_fastest": safe_median(time_gaps),
                "mean_time_gap_to_fastest": safe_mean(time_gaps),
                "median_wall_time_ms": safe_median(wall_times_ms),
                "p90_wall_time_ms": percentile(wall_times_ms, 0.90),
                "median_wall_time_per_customer_us": safe_median(wall_per_customer_us),
                "log_error_runs": sum(1 for row in rows if row.get("log_has_error")),
            }
        )

    return output


def build_baseline_matrix(run_rows: Sequence[dict[str, Any]], algorithms: Sequence[str], context: SummaryContext) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in run_rows:
        grouped[str(row["instance_name"])].append(row)

    output: list[dict[str, Any]] = []
    ordered_algorithms = list(sorted(set(algorithms)))

    for instance_name, rows in sorted(grouped.items(), key=lambda item: ((as_int(item[1][0].get("customer_count")) or -1), item[0])):
        base: dict[str, Any] = {
            "instance_name": instance_name,
            "instance_type": as_str(rows[0].get("instance_type")),
            "suite_names": ",".join(sorted({str(row["suite_name"]) for row in rows if row.get("suite_name")})),
            "depot_count": as_int(rows[0].get("depot_count")),
            "customer_count": as_int(rows[0].get("customer_count")),
            "salesman_count": as_int(rows[0].get("salesman_count")),
            "return_to_depot": as_bool(rows[0].get("return_to_depot")),
            "best_observed_objective": context.best_objective_by_instance.get(instance_name),
            "fastest_observed_wall_time_ms": context.fastest_wall_time_ms_by_instance.get(instance_name),
        }

        by_algorithm: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            by_algorithm[str(row["algorithm_id"])].append(row)

        for algorithm_id in ordered_algorithms:
            algo_rows = by_algorithm.get(algorithm_id, [])
            feasible_rows = [row for row in algo_rows if row.get("objective") is not None and row.get("is_feasible_success")]
            runtime_rows = [row for row in algo_rows if row.get("wall_time_ms") is not None and row.get("is_success")]
            objectives = [as_float(row.get("objective")) for row in feasible_rows if as_float(row.get("objective")) is not None]
            runtimes = [as_float(row.get("wall_time_ms")) for row in runtime_rows if as_float(row.get("wall_time_ms")) is not None]
            gaps = [as_float(row.get("gap_to_best_observed")) for row in feasible_rows if as_float(row.get("gap_to_best_observed")) is not None]

            prefix = f"{algorithm_id}__"
            base[prefix + "runs"] = len(algo_rows)
            base[prefix + "feasible_rate"] = (sum(1 for row in algo_rows if row.get("is_feasible_success")) / len(algo_rows) if algo_rows else None)
            base[prefix + "best_objective"] = min(objectives) if objectives else None
            base[prefix + "median_objective"] = safe_median(objectives)
            base[prefix + "median_gap_to_best_observed"] = safe_median(gaps)
            base[prefix + "median_wall_time_ms"] = safe_median(runtimes)
            base[prefix + "log_error_runs"] = sum(1 for row in algo_rows if row.get("log_has_error"))

        output.append(base)

    return output


def build_failures(run_rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    failures: list[dict[str, Any]] = []
    for row in run_rows:
        if not row.get("is_flagged_failure"):
            continue
        failures.append(
            {
                "run_id": row.get("run_id"),
                "suite_name": row.get("suite_name"),
                "algorithm_id": row.get("algorithm_id"),
                "instance_name": row.get("instance_name"),
                "seed": row.get("seed"),
                "status": row.get("status"),
                "feasible": row.get("feasible"),
                "objective": row.get("objective"),
                "wall_time_ms": row.get("wall_time_ms"),
                "log_has_error": row.get("log_has_error"),
                "log_has_warning": row.get("log_has_warning"),
                "log_first_error_line": row.get("log_first_error_line"),
                "log_tail_excerpt": row.get("log_tail_excerpt"),
                "result_file": row.get("relative_result_file") or row.get("result_file"),
                "log_file": row.get("relative_log_file") or row.get("log_file"),
            }
        )
    failures.sort(key=lambda row: (str(row.get("algorithm_id") or ""), str(row.get("instance_name") or ""), as_int(row.get("seed")) or -1))
    return failures


def build_export_bundle(
    runs_roots: Sequence[Path],
    logs_roots: Sequence[Path],
    repo_root: Path | None,
    *,
    max_tail_lines: int,
    max_cell_text: int,
) -> ExportBundle:
    run_rows, run_issues = collect_runs(runs_roots, repo_root)
    log_records, log_issues = collect_logs(
        logs_roots,
        repo_root,
        max_tail_lines=max_tail_lines,
        max_cell_text=max_cell_text,
    )

    attach_logs(run_rows, log_records, repo_root)
    context = build_summary_context(run_rows)
    enrich_runs_with_relative_metrics(run_rows, context)

    algorithms = summarize_algorithms(run_rows, context)
    instances = summarize_instances(run_rows, context)
    algorithm_instance = summarize_algorithm_instance(run_rows, context)
    algorithm_instance_type = summarize_algorithm_instance_type(run_rows)
    baseline_matrix = build_baseline_matrix(run_rows, [row["algorithm_id"] for row in run_rows], context)
    failures = build_failures(run_rows)
    logs = [record.as_row() for _, record in sorted(log_records.items())]
    parse_issues = [asdict(issue) for issue in (*run_issues, *log_issues)]

    return ExportBundle(
        runs=run_rows,
        algorithms=algorithms,
        instances=instances,
        algorithm_instance=algorithm_instance,
        algorithm_instance_type=algorithm_instance_type,
        baseline_matrix=baseline_matrix,
        failures=failures,
        logs=logs,
        parse_issues=parse_issues,
    )


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", title).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:MAX_SHEET_NAME_LEN]


def write_table_sheet(
    wb: Workbook,
    *,
    title: str,
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str] | None = None,
    freeze_cell: str = "A2",
    percent_columns: set[str] | None = None,
    integer_columns: set[str] | None = None,
    float_columns: set[str] | None = None,
    wrap_columns: set[str] | None = None,
) -> Any:
    ws = wb.create_sheet(title=safe_sheet_title(title))
    percent_columns = percent_columns or set()
    integer_columns = integer_columns or set()
    float_columns = float_columns or set()
    wrap_columns = wrap_columns or set()

    if rows:
        if columns is None:
            columns = list(rows[0].keys())
    else:
        columns = list(columns or ["message"])

    ws.append(list(columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in rows:
        values = [normalize_cell_value(row.get(column)) for column in columns]
        ws.append(values)

    if not rows:
        ws.append(["No data"] + [None] * (len(columns) - 1))

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column in wrap_columns)
            if column in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
            elif column in integer_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0"
            elif column in float_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"

    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    if ws.max_row >= 2 and ws.max_column >= 1:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)
        table = Table(displayName=f"tbl_{table_name}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    if "feasible_rate" in columns:
        col = get_column_letter(columns.index("feasible_rate") + 1)
        if ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=0.5, mid_color="FFEB84", end_type="num", end_value=1, end_color="63BE7B"),
            )

    for maybe_gap in ["median_gap_to_best_observed", "mean_gap_to_best_observed", "gap_to_best_observed", "median_time_gap_to_fastest", "time_gap_to_fastest"]:
        if maybe_gap in columns and ws.max_row >= 2:
            col = get_column_letter(columns.index(maybe_gap) + 1)
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B", mid_type="num", mid_value=0.25, mid_color="FFEB84", end_type="num", end_value=1, end_color="F8696B"),
            )

    if "log_has_error" in columns and ws.max_row >= 2:
        col = get_column_letter(columns.index("log_has_error") + 1)
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(operator="equal", formula=["TRUE"], fill=FAIL_FILL),
        )

    apply_column_widths(ws, wrap_columns=wrap_columns)
    add_hyperlinks(ws, columns)
    return ws


def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return None
    text = str(value)
    return trim_text(text, DEFAULT_MAX_CELL_TEXT)


def apply_column_widths(ws: Any, *, wrap_columns: set[str]) -> None:
    header_map = {cell.column: str(cell.value) if cell.value is not None else "" for cell in ws[1]}
    for col_idx in range(1, ws.max_column + 1):
        header = header_map.get(col_idx, "")
        max_len = len(header)
        scan_limit = min(ws.max_row, 250)
        for row_idx in range(2, scan_limit + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = min(len(text), 80)
        width = max(10, min(40 if header in wrap_columns else 24, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_hyperlinks(ws: Any, columns: Sequence[str]) -> None:
    file_cols = {idx + 1 for idx, column in enumerate(columns) if column.endswith("_file") or column.endswith("_path")}
    if not file_cols:
        return
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in file_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.value:
                continue
            text = str(cell.value)
            if text.startswith("/"):
                cell.hyperlink = Path(text).as_uri()
                cell.style = "Hyperlink"


def build_overview_sheet(wb: Workbook, bundle: ExportBundle, report_path: Path, repo_root: Path | None) -> None:
    ws = wb.create_sheet(title="Overview", index=0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "MDMTSP Research Excel Report"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:D1")

    metrics = {
        "Generated file": str(report_path),
        "Repository root": str(repo_root) if repo_root else "n/a",
        "Runs": len(bundle.runs),
        "Algorithms": len(bundle.algorithms),
        "Instances": len(bundle.instances),
        "Failures": len(bundle.failures),
        "Parse issues": len(bundle.parse_issues),
        "Logs": len(bundle.logs),
    }

    row = 3
    for key, value in metrics.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    start_row = 3
    chart_anchor_row = max(3, row + 1)

    algo_header_row = chart_anchor_row
    ws.cell(row=algo_header_row, column=4, value="Algorithm")
    ws.cell(row=algo_header_row, column=5, value="Median gap")
    ws.cell(row=algo_header_row, column=6, value="Median wall time ms")
    for col in range(4, 7):
        cell = ws.cell(row=algo_header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for idx, row_data in enumerate(bundle.algorithms, start=algo_header_row + 1):
        ws.cell(row=idx, column=4, value=row_data.get("algorithm_id"))
        ws.cell(row=idx, column=5, value=row_data.get("median_gap_to_best_observed"))
        ws.cell(row=idx, column=6, value=row_data.get("median_wall_time_ms"))
        ws.cell(row=idx, column=5).number_format = "0.00%"
        ws.cell(row=idx, column=6).number_format = "0.000"
        for col in range(4, 7):
            ws.cell(row=idx, column=col).border = THIN_BORDER

    if bundle.algorithms:
        gap_chart = BarChart()
        gap_chart.title = "Median gap to best observed"
        gap_chart.y_axis.title = "Gap"
        gap_chart.x_axis.title = "Algorithm"
        gap_data = Reference(ws, min_col=5, min_row=algo_header_row, max_row=algo_header_row + len(bundle.algorithms))
        gap_labels = Reference(ws, min_col=4, min_row=algo_header_row + 1, max_row=algo_header_row + len(bundle.algorithms))
        gap_chart.add_data(gap_data, titles_from_data=True)
        gap_chart.set_categories(gap_labels)
        gap_chart.height = 7
        gap_chart.width = 11
        ws.add_chart(gap_chart, "A13")

        time_chart = BarChart()
        time_chart.title = "Median wall time (ms)"
        time_chart.y_axis.title = "ms"
        time_chart.x_axis.title = "Algorithm"
        time_data = Reference(ws, min_col=6, min_row=algo_header_row, max_row=algo_header_row + len(bundle.algorithms))
        time_labels = Reference(ws, min_col=4, min_row=algo_header_row + 1, max_row=algo_header_row + len(bundle.algorithms))
        time_chart.add_data(time_data, titles_from_data=True)
        time_chart.set_categories(time_labels)
        time_chart.height = 7
        time_chart.width = 11
        ws.add_chart(time_chart, "L13")

    definitions_row = max(13, algo_header_row + len(bundle.algorithms) + 3)
    ws.cell(row=definitions_row, column=1, value="Key definitions")
    ws.cell(row=definitions_row, column=1).fill = HEADER_FILL
    ws.cell(row=definitions_row, column=1).font = HEADER_FONT
    ws.merge_cells(start_row=definitions_row, start_column=1, end_row=definitions_row, end_column=4)

    definitions = [
        ("gap_to_best_observed", "(objective - best_observed_objective) / best_observed_objective"),
        ("time_ratio_to_fastest", "wall_time_ms / fastest_observed_wall_time_ms"),
        ("time_gap_to_fastest", "(wall_time_ms - fastest_observed_wall_time_ms) / fastest_observed_wall_time_ms"),
        ("is_feasible_success", "status in {ok, success} and feasible is not false"),
    ]
    for idx, (name, formula) in enumerate(definitions, start=definitions_row + 1):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=formula)
        ws.cell(row=idx, column=1).fill = SUBHEADER_FILL
        ws.cell(row=idx, column=1).font = Font(bold=True)
        ws.cell(row=idx, column=1).border = THIN_BORDER
        ws.cell(row=idx, column=2).border = THIN_BORDER
        ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True)

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28 if col in {1, 2} else 18


def build_definitions_sheet(wb: Workbook) -> None:
    rows = [
        {
            "field": "run_id",
            "meaning": "Unique identifier of a single solver run.",
        },
        {
            "field": "feasible_rate",
            "meaning": "feasible_runs / runs inside the corresponding grouping.",
        },
        {
            "field": "gap_to_best_observed",
            "meaning": "(objective - best_observed_objective) / best_observed_objective.",
        },
        {
            "field": "time_ratio_to_fastest",
            "meaning": "wall_time_ms / fastest_observed_wall_time_ms for the same instance.",
        },
        {
            "field": "time_gap_to_fastest",
            "meaning": "(wall_time_ms - fastest_observed_wall_time_ms) / fastest_observed_wall_time_ms.",
        },
        {
            "field": "log_has_error",
            "meaning": "Best-effort heuristic flag based on stderr / log keywords such as error, failed, exception, traceback.",
        },
        {
            "field": "baseline_matrix",
            "meaning": "Wide-format comparison table with per-instance metrics for every algorithm.",
        },
    ]
    write_table_sheet(
        wb,
        title="Definitions",
        rows=rows,
        columns=["field", "meaning"],
        wrap_columns={"meaning"},
    )


def write_workbook(bundle: ExportBundle, output_path: Path, repo_root: Path | None) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    build_overview_sheet(wb, bundle, output_path, repo_root)
    build_definitions_sheet(wb)

    percent_cols = {
        "feasible_rate",
        "gap_to_best_observed",
        "time_gap_to_fastest",
        "min_gap_to_best_observed",
        "mean_gap_to_best_observed",
        "median_gap_to_best_observed",
        "min_time_gap_to_fastest",
        "mean_time_gap_to_fastest",
        "median_time_gap_to_fastest",
        "median_gap_to_best_observed",
        "mean_gap_to_best_observed",
        "median_time_gap_to_fastest",
        "mean_time_gap_to_fastest",
    }
    integer_cols = {
        "runs",
        "successful_runs",
        "feasible_runs",
        "objective_comparable_runs",
        "time_comparable_runs",
        "unique_instances",
        "unique_instance_types",
        "unique_suites",
        "best_run_count",
        "best_instance_coverage",
        "fastest_run_count",
        "fastest_instance_coverage",
        "depot_count",
        "customer_count",
        "salesman_count",
        "route_count",
        "seed",
        "log_warning_count",
        "log_error_count",
        "log_traceback_count",
        "log_error_runs",
        "log_warning_runs",
        "algorithms_tested",
        "seeds_tested",
        "size_bytes",
        "line_count",
    }
    float_cols = {
        "objective",
        "best_objective",
        "mean_objective",
        "median_objective",
        "wall_time_ms",
        "wall_time_s",
        "wall_time_per_customer_us",
        "min_time_ratio_to_fastest",
        "mean_time_ratio_to_fastest",
        "median_time_ratio_to_fastest",
        "total_wall_time_s",
        "mean_wall_time_ms",
        "median_wall_time_ms",
        "p90_wall_time_ms",
        "median_wall_time_per_customer_us",
        "best_observed_objective",
        "fastest_observed_wall_time_ms",
        "median_wall_time_ms",
        "median_wall_time_per_customer_us",
    }
    wrap_cols = {
        "instance_types",
        "suite_names",
        "log_first_error_line",
        "log_first_warning_line",
        "log_last_nonempty_line",
        "log_tail_excerpt",
        "tail_excerpt",
        "first_error_line",
        "first_warning_line",
        "last_nonempty_line",
        "reason",
        "meaning",
    }

    write_table_sheet(
        wb,
        title="Runs",
        rows=bundle.runs,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Algorithms",
        rows=bundle.algorithms,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Instances",
        rows=bundle.instances,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Algo_Instance",
        rows=bundle.algorithm_instance,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Algo_Type",
        rows=bundle.algorithm_instance_type,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Baseline_Matrix",
        rows=bundle.baseline_matrix,
        percent_columns={column for column in (bundle.baseline_matrix[0].keys() if bundle.baseline_matrix else []) if "feasible_rate" in column or "gap" in column},
        integer_columns={column for column in (bundle.baseline_matrix[0].keys() if bundle.baseline_matrix else []) if column.endswith("__runs") or column.endswith("__log_error_runs")}.union({"depot_count", "customer_count", "salesman_count"}),
        float_columns={column for column in (bundle.baseline_matrix[0].keys() if bundle.baseline_matrix else []) if column.endswith("__best_objective") or column.endswith("__median_objective") or column.endswith("__median_wall_time_ms")}.union({"best_observed_objective", "fastest_observed_wall_time_ms"}),
        wrap_columns={"suite_names"},
    )
    write_table_sheet(
        wb,
        title="Failures",
        rows=bundle.failures,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Log_Diagnostics",
        rows=bundle.logs,
        percent_columns=percent_cols,
        integer_columns=integer_cols,
        float_columns=float_cols,
        wrap_columns=wrap_cols,
    )
    write_table_sheet(
        wb,
        title="Parse_Issues",
        rows=bundle.parse_issues,
        columns=["source_kind", "path", "reason"],
        wrap_columns={"reason", "path"},
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(output_path.parent)) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        tmp_path.replace(output_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def main() -> int:
    args = parse_args()
    repo_root = args.repo_root.resolve() if args.repo_root else repo_root_from_script()
    runs_roots = resolve_paths(args.runs_roots, repo_root / "results" / "runs")
    logs_roots = resolve_paths(args.logs_roots, repo_root / "results" / "logs")
    output_path = args.output.resolve() if args.output.is_absolute() else (repo_root / args.output).resolve()

    if args.max_tail_lines <= 0:
        raise SystemExit("--max-tail-lines must be positive")
    if args.max_cell_text <= 0:
        raise SystemExit("--max-cell-text must be positive")

    bundle = build_export_bundle(
        runs_roots,
        logs_roots,
        repo_root,
        max_tail_lines=args.max_tail_lines,
        max_cell_text=min(args.max_cell_text, DEFAULT_MAX_CELL_TEXT),
    )

    write_workbook(bundle, output_path, repo_root)
    print(f"excel report saved to {output_path}")
    print(
        f"runs={len(bundle.runs)} algorithms={len(bundle.algorithms)} instances={len(bundle.instances)} failures={len(bundle.failures)} parse_issues={len(bundle.parse_issues)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
